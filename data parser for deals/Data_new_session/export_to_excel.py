import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import argparse

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'goszakup_2024_2026.db')

COLUMNS = [
    ("contract_id",      "ID договора"),
    ("contract_number",  "Номер договора"),
    ("purchase_number",  "Номер закупки"),
    ("contract_type",    "Тип договора"),
    ("contract_date",    "Дата договора"),
    ("contract_status",  "Статус договора"),
    ("purchase_method",  "Способ закупки"),
    ("customer_name",    "Заказчик"),
    ("customer_bin",     "БИН заказчика"),
    ("supplier_name",    "Поставщик"),
    ("supplier_bin",     "БИН поставщика"),
    ("lot_id",           "ID лота"),
    ("lot_title",        "Наименование товара"),
    ("unit_price",       "Цена за единицу"),
    ("quantity",         "Количество"),
    ("contract_amount",  "Сумма лота"),
    ("brand_model",      "Марка/модель"),
    ("country",          "Страна"),
    ("manufacturer",     "Производитель"),
]

COL_WIDTHS = [15, 22, 16, 20, 18, 18, 25, 45, 14, 45, 14, 15, 60, 16, 12, 18, 25, 18, 35]

def export_db(db_path: str = DB_PATH, output_file: str = None):
    if not output_file:
        output_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', f"goszakup_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        
    print(f"Connecting to {db_path}...")
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    
    db_cols = [c[0] for c in COLUMNS]
    cur.execute(f"SELECT {', '.join(db_cols)} FROM contracts_lots ORDER BY contract_date, contract_number")
    
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Договоры Товары 2024-2026")
    
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    CENTER = Alignment(horizontal="center", vertical="center")
    
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        
    header_cells = []
    for label in [c[1] for c in COLUMNS]:
        cell = openpyxl.cell.WriteOnlyCell(ws, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        header_cells.append(cell)
    ws.append(header_cells)
    
    BATCH = 10000
    total = 0
    while True:
        rows = cur.fetchmany(BATCH)
        if not rows:
            break
        for row in rows:
            ws.append(list(row))
            total += 1
        print(f"  Exported rows: {total:,}", end="\r", flush=True)
        
    conn.close()
    
    ws2 = wb.create_sheet("Сводка")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 55
    summary = [
        ("Дата выгрузки",         datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("Способы закупки",       "ЗЦП, ОК, ОИ (все)"),
        ("Вид предмета закупки",   "Товар"),
        ("Статусы",                "Действует, Изменен, Исполнен, Подписан, Создано доп.соглашение, Частично исполнен"),
        ("Период",                 "01.01.2024 – 25.08.2026"),
        ("Всего лотов",            total),
    ]
    for k, v in summary:
        ws2.append([k, v])
        
    print(f"\nTotal rows exported: {total:,}")
    print(f"Saving {output_file} ...")
    wb.save(output_file)
    print(f"✅ Export completed: {output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", type=str, default=DB_PATH)
    parser.add_argument("--out", type=str, default=None)
    args = parser.parse_args()
    export_db(args.db, args.out)
