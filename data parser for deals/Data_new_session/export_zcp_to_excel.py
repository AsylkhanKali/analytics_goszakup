import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

DB_PATH = "data/goszakup.db"
OUTPUT_FILE = f"data/contracts_zcp_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

COLUMNS = [
    ("contract_id",      "ID договора"),
    ("contract_number",  "Номер договора"),
    ("contract_date",    "Дата договора"),
    ("contract_status",  "Статус"),
    ("customer_name",    "Заказчик"),
    ("customer_bin",     "БИН заказчика"),
    ("supplier_name",    "Поставщик"),
    ("supplier_bin",     "БИН поставщика"),
    ("lot_title",        "Наименование товара"),
    ("unit_price",       "Цена за единицу"),
    ("quantity",         "Количество"),
    ("contract_amount",  "Сумма договора"),
    ("brand_model",      "Марка/модель"),
    ("country",          "Страна"),
    ("manufacturer",     "Производитель"),
]

COL_WIDTHS = [18, 22, 14, 18, 45, 14, 45, 14, 60, 16, 12, 18, 25, 18, 35]

def main():
    print(f"Подключаемся к {DB_PATH}...")
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()

    db_cols = [c[0] for c in COLUMNS]
    cur.execute(f"SELECT {', '.join(db_cols)} FROM contracts_zcp ORDER BY contract_date, contract_number")

    wb = openpyxl.Workbook(write_only=True)   # write_only = намного быстрее
    ws = wb.create_sheet("ЗЦП Товары 2024-2026")

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    CENTER = Alignment(horizontal="center", vertical="center")

    # Задаём ширины колонок
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Заголовок
    header_cells = []
    for label in [c[1] for c in COLUMNS]:
        cell = openpyxl.cell.WriteOnlyCell(ws, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        header_cells.append(cell)
    ws.append(header_cells)

    # Данные — без побочного форматирования, только значения
    BATCH = 10000
    total = 0
    while True:
        rows = cur.fetchmany(BATCH)
        if not rows:
            break
        for row in rows:
            ws.append(list(row))
            total += 1
        print(f"  Записано строк: {total:,}", end="\r", flush=True)

    con.close()

    # Лист сводки
    ws2 = wb.create_sheet("Сводка")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 55
    summary = [
        ("Дата выгрузки",         datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("Способ закупки",         "ЗЦП (метод 3)"),
        ("Вид предмета закупки",   "Товар"),
        ("Статусы",                "Исполнен, Подписан, Действует, Изменен, Частично исполнен"),
        ("Период",                 "01.01.2024 – 31.01.2026"),
        ("Всего лотов",            total),
    ]
    for k, v in summary:
        ws2.append([k, v])

    print(f"\nВсего строк: {total:,}")
    print(f"Сохраняем {OUTPUT_FILE} ...")
    wb.save(OUTPUT_FILE)
    print(f"✅ Готово: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
